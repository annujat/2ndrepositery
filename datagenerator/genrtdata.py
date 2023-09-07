import openpyxl

def datadrivens():
    wb=openpyxl.load_workbook("C:/Users/Aniket/Desktop/xl.xlsx")
    ws=wb['Sheet']
    r=ws.max_row
    list1 = []
    list2 = []
    for i in range(1,r+1):
        list1=[]
        un=ws.cell(i,1)
        up=ws.cell(i,2)
        list1.insert(0,un.value)
        list1.insert(1,up.value)
        list2.insert(i-1,list1)
    print(list2)
    return list2




